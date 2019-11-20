#---------How to run
#python ANSYS-dat_to_vtm.py mesh.dat Table.xlsx
import sys
Dir=''						# Work directory
MeshFileName=sys.argv[1]	# ANSYS dat file
BCsFileName=sys.argv[2]		# Excel file with thermal boundary conditions
#---------Example---------------------
#Dir='D:\\Ansaldo\\GTI\\Annotations\\'
#MeshFileName='mesh.dat'
#BCsFileName='Table.xlsx'
#--Settings--------------------------------
print('*Excel-file reading')
Hex=[[0,1,2,3],[0,1,5,4],[1,2,6,5],[2,3,7,6],[0,4,7,3],[4,5,6,7]]
Tet=[[0,1,2],[0,1,3],[1,2,3],[0,3,2]]
#path to Excel reading module
#sys.path.append('C:\\Users\\alexey.makrushin\\AppData\\Local\\Continuum\\anaconda2\\Lib\\site-packages')
#path to vtk module
sys.path.append('D:\\Programs\\ParaView-5.7.0-RC1-Windows-msvc2015-64bit\\bin\\Lib\\site-packages')
import vtk
#------------------------------------------
from openpyxl import load_workbook
wb=load_workbook(Dir+BCsFileName, read_only=True)
ws=wb['BC Zones']
TableRows={}
for i in range(4,ws.max_row+1): TableRows[str(ws.cell(i,15).value).upper()]=i
#----------------------------------------------------------
#------------DAT-file reading------------------------------
print('*Mesh reading')
ElemNodesNum=[0,0,0,0,0,0,0,0]
PointsArray={}
#=vtk.vtkPoints()
#PointsArray.InsertNextPoint(0.0,0.0,0.0)
Faces={}	#Keys:[Min][Max][Third]
CMBLOCK={}
BlocksList=[]
BlocksList.append('')
BlocksInfo={}
BlocksInfo['']=[0,0,vtk.vtkUnstructuredGrid(),{},0,vtk.vtkPoints()] 	# Index in the List, Number of Cells, vtkUnstructuredGrid, Points, Number of Points, vtk.vtkPoints()
mf=open(Dir+MeshFileName,'r')
txt=mf.readline()
while txt[0:5]!='/solu':
	if txt[0:25]=='/com,*********** Create "':
		cmnt=txt.split('"')[1].upper()
		NamedSelection=''
		for TableKey in TableRows:
			if (TableKey in cmnt)and(not TableKey in BlocksInfo):
				NamedSelection=TableKey
				BlocksInfo[NamedSelection]=[len(BlocksList),0,vtk.vtkUnstructuredGrid(),{},0,vtk.vtkPoints()]
				BlocksList.append(NamedSelection)
				print(NamedSelection+' has been found')
	elif txt[0:3]=='et,':
		ElemType=int(txt.split(',')[2])
		if ElemType==186 or ElemType==187: print('***ERROR - There are elemets for structural calculations!!!!!!')
	elif txt[0:7]=='nblock,':			#-----Nodes block
		Num=int(txt.split(',')[3])
		txt=mf.readline()
		Values=txt[1:-2].split(',')
		NumLen=int(Values[0].split('i')[1])
		RowNum=int(Values[0].split('i')[0])		
		CoordLen=int(Values[1].split('.')[0].split('e')[1])
		for i in range(0,Num):
			txt=mf.readline()
			TxtLen=len(txt)-1
			if TxtLen>(RowNum*NumLen+2*CoordLen): PointsArray[int(txt[0:NumLen])]=(float(txt[NumLen:NumLen+CoordLen]),float(txt[NumLen+CoordLen:NumLen+2*CoordLen]),float(txt[NumLen+2*CoordLen:NumLen+3*CoordLen]))
			elif TxtLen>(RowNum*NumLen+CoordLen): PointsArray[int(txt[0:NumLen])]=(float(txt[NumLen:NumLen+CoordLen]),float(txt[NumLen+CoordLen:NumLen+2*CoordLen]),0.0)
			else: PointsArray[int(txt[0:NumLen])]=(float(txt[NumLen:NumLen+CoordLen]),0.0,0.0)
	elif txt[0:7]=='eblock,':
		NumLen=int(mf.readline()[1:-2].split('i')[1])
		txt=mf.readline()
		if ElemType==87:		#-----QUADRATIC TETRA
			while txt[0:2]!='-1':
				for i in range(0,4):ElemNodesNum[i]=int(txt[(i+11)*NumLen:(i+12)*NumLen])
				for Face in Tet:
					FaceNodesNum=[ElemNodesNum[Face[0]],ElemNodesNum[Face[1]],ElemNodesNum[Face[2]]]
					MinNum=FaceNodesNum[0]
					MaxNum1=FaceNodesNum[0]
					if MinNum>FaceNodesNum[1]:
						MaxNum1=MinNum
						MinNum=FaceNodesNum[1]
					else:MaxNum1=FaceNodesNum[1]
					if MinNum>FaceNodesNum[2]:
						MidNum1=MinNum
						MinNum=FaceNodesNum[2]
					elif MaxNum1<FaceNodesNum[2]:
						MidNum1=MaxNum1
						MaxNum1=FaceNodesNum[2]
					else:MidNum1=FaceNodesNum[2]
					if not MinNum in Faces: Faces[MinNum]={}
					if not MaxNum1 in Faces[MinNum]: Faces[MinNum][MaxNum1]={}
					if not MidNum1 in Faces[MinNum][MaxNum1]: Faces[MinNum][MaxNum1][MidNum1]=0
					else:
						Faces[MinNum][MaxNum1].__delitem__(MidNum1)
						if len(Faces[MinNum][MaxNum1])==0:
							Faces[MinNum].__delitem__(MaxNum1)
							if len(Faces[MinNum])==0:Faces.__delitem__(MinNum)
				txt=mf.readline()
				txt=mf.readline()
		elif ElemType==90:		#-----QUADRATIC HEX
			while txt[0:2]!='-1':
				for i in range(0,8):ElemNodesNum[i]=int(txt[(i+11)*NumLen:(i+12)*NumLen])
				for Face in Hex:
					FaceNodesNum=[ElemNodesNum[Face[0]],ElemNodesNum[Face[1]],ElemNodesNum[Face[2]],ElemNodesNum[Face[3]]]
					MinNum=int(min(FaceNodesNum))
					indx=FaceNodesNum.index(MinNum)
					if indx==3:
						MaxNum1=FaceNodesNum[0]
						MidNum1=FaceNodesNum[0]
					else:
						MaxNum1=FaceNodesNum[indx+1]
						MidNum1=FaceNodesNum[indx+1]
					if indx==0:
						MaxNum2=FaceNodesNum[3]
						MidNum2=FaceNodesNum[3]
					else:
						MaxNum2=FaceNodesNum[indx-1]
						MidNum2=FaceNodesNum[indx-1]
					#----------
					if indx>1:
						if MaxNum1<FaceNodesNum[indx-2]:MaxNum1=FaceNodesNum[indx-2]
						else:MidNum1=FaceNodesNum[indx-2]
						if MaxNum2<FaceNodesNum[indx-2]:MaxNum2=FaceNodesNum[indx-2]
						else:MidNum2=FaceNodesNum[indx-2]					
					else:
						if MaxNum1<FaceNodesNum[indx+2]:MaxNum1=FaceNodesNum[indx+2]
						else:MidNum1=FaceNodesNum[indx+2]
						if MaxNum2<FaceNodesNum[indx+2]:MaxNum2=FaceNodesNum[indx+2]
						else:MidNum2=FaceNodesNum[indx+2]
					if not MinNum in Faces: Faces[MinNum]={}
					if not MaxNum1 in Faces[MinNum]: Faces[MinNum][MaxNum1]={}
					if not MaxNum2 in Faces[MinNum]: Faces[MinNum][MaxNum2]={}
					if not MidNum1 in Faces[MinNum][MaxNum1]: Faces[MinNum][MaxNum1][MidNum1]=0
					else:
						Faces[MinNum][MaxNum1].__delitem__(MidNum1)
						if len(Faces[MinNum][MaxNum1])==0:
							Faces[MinNum].__delitem__(MaxNum1)
							if len(Faces[MinNum])==0:Faces.__delitem__(MinNum)
					if not MidNum2 in Faces[MinNum][MaxNum2]: Faces[MinNum][MaxNum2][MidNum2]=0
					else:
						Faces[MinNum][MaxNum2].__delitem__(MidNum2)
						if len(Faces[MinNum][MaxNum2])==0:
							Faces[MinNum].__delitem__(MaxNum2)
							if len(Faces[MinNum])==0:Faces.__delitem__(MinNum)
				txt=mf.readline()
				txt=mf.readline()
		elif ElemType==152 and NamedSelection!='':		#-----SURFACE ELEMENTS
			while txt[0:2]!='-1':
				for i in range(0,4):ElemNodesNum[i]=int(txt[(i+5)*NumLen:(i+6)*NumLen])
				if ElemNodesNum[2]==ElemNodesNum[3]:
					FaceNodesNum=[ElemNodesNum[0],ElemNodesNum[1],ElemNodesNum[2]]
					MinNum=int(min(FaceNodesNum))
					FaceNodesNum.remove(MinNum)
					MaxNum1=int(max(FaceNodesNum))
					FaceNodesNum.remove(MaxNum1)
					if MinNum in Faces:
						if MaxNum1 in Faces[MinNum]:
							if FaceNodesNum[0] in Faces[MinNum][MaxNum1]:Faces[MinNum][MaxNum1][FaceNodesNum[0]]=BlocksInfo[NamedSelection][0]
				else:
					FaceNodesNum=[ElemNodesNum[0],ElemNodesNum[1],ElemNodesNum[2],ElemNodesNum[3]]
					MinNum=int(min(FaceNodesNum))
					indx=FaceNodesNum.index(MinNum)
					if indx==3:
						MaxNum1=FaceNodesNum[0]
						MidNum1=FaceNodesNum[0]
					else:
						MaxNum1=FaceNodesNum[indx+1]
						MidNum1=FaceNodesNum[indx+1]
					if indx==0:
						MaxNum2=FaceNodesNum[3]
						MidNum2=FaceNodesNum[3]
					else:
						MaxNum2=FaceNodesNum[indx-1]
						MidNum2=FaceNodesNum[indx-1]
					#----------
					if indx>1:
						if MaxNum1<FaceNodesNum[indx-2]:MaxNum1=FaceNodesNum[indx-2]
						else:MidNum1=FaceNodesNum[indx-2]
						if MaxNum2<FaceNodesNum[indx-2]:MaxNum2=FaceNodesNum[indx-2]
						else:MidNum2=FaceNodesNum[indx-2]					
					else:
						if MaxNum1<FaceNodesNum[indx+2]:MaxNum1=FaceNodesNum[indx+2]
						else:MidNum1=FaceNodesNum[indx+2]
						if MaxNum2<FaceNodesNum[indx+2]:MaxNum2=FaceNodesNum[indx+2]
						else:MidNum2=FaceNodesNum[indx+2]
					if MinNum in Faces:
						if MaxNum1 in Faces[MinNum]:
							if MidNum1 in Faces[MinNum][MaxNum1]: Faces[MinNum][MaxNum1][MidNum1]=BlocksInfo[NamedSelection][0]						
						if MaxNum2 in Faces[MinNum]:
							if MidNum2 in Faces[MinNum][MaxNum2]: Faces[MinNum][MaxNum2][MidNum2]=BlocksInfo[NamedSelection][0]
				txt=mf.readline()
	elif txt[0:8]=='CMBLOCK,':
		Values=txt.split(',')
		NamedSelection=Values[1].replace(' ','').upper()
		NodesNum=int(Values[3])
		if NamedSelection in TableRows:
			CMBLOCK[NamedSelection]=[]
			BlocksInfo[NamedSelection]=[len(BlocksList),0,vtk.vtkUnstructuredGrid(),{},0,vtk.vtkPoints()]
			BlocksList.append(NamedSelection)
			print(NamedSelection+' has been found')
			NumLen=int(mf.readline()[1:-2].split('i')[1])
			j=0
			while j<NodesNum:
				txt=mf.readline()
				k0=len(txt)-1
				k=0					
				while k*NumLen<k0:
					CMBLOCK[NamedSelection].append(int(txt[k*NumLen:(k+1)*NumLen]))
					k+=1
				j+=k
	txt=mf.readline()			
mf.close()
#----------------------------------------------------------
#------------DATA processing-------------------------------
print('**Data processing')
#-- Node blocks and cells counter
for MinNum in Faces:
	for MaxNum1 in Faces[MinNum]:
		for MidNum1 in Faces[MinNum][MaxNum1]:
			for NamedSelection in CMBLOCK:
				if (MinNum in CMBLOCK[NamedSelection])and(MaxNum1 in CMBLOCK[NamedSelection])and(MidNum1 in CMBLOCK[NamedSelection]):
					Faces[MinNum][MaxNum1][MidNum1]=BlocksInfo[NamedSelection][0]
			NamedSelection=BlocksList[Faces[MinNum][MaxNum1][MidNum1]]
			BlocksInfo[NamedSelection][1]+=1
			if not MinNum in BlocksInfo[NamedSelection][3]:
				BlocksInfo[NamedSelection][3][MinNum]=BlocksInfo[NamedSelection][4]
				BlocksInfo[NamedSelection][4]+=1
				BlocksInfo[NamedSelection][5].InsertNextPoint(PointsArray[MinNum][0],PointsArray[MinNum][1],PointsArray[MinNum][2])
			if not MidNum1 in BlocksInfo[NamedSelection][3]:
				BlocksInfo[NamedSelection][3][MidNum1]=BlocksInfo[NamedSelection][4]
				BlocksInfo[NamedSelection][4]+=1
				BlocksInfo[NamedSelection][5].InsertNextPoint(PointsArray[MidNum1][0],PointsArray[MidNum1][1],PointsArray[MidNum1][2])
			if not MaxNum1 in BlocksInfo[NamedSelection][3]:
				BlocksInfo[NamedSelection][3][MaxNum1]=BlocksInfo[NamedSelection][4]
				BlocksInfo[NamedSelection][4]+=1
				BlocksInfo[NamedSelection][5].InsertNextPoint(PointsArray[MaxNum1][0],PointsArray[MaxNum1][1],PointsArray[MaxNum1][2])
#-- Blocks preaparation
for NamedSelection in BlocksList:
	if BlocksInfo[NamedSelection][1]==0:
		print('*WARNING: '+NamedSelection+' is overlapping and removed')
		BlocksInfo.__delitem__(NamedSelection)
	else:
		BlocksInfo[NamedSelection][2].SetPoints(BlocksInfo[NamedSelection][5])
		BlocksInfo[NamedSelection][2].Allocate(BlocksInfo[NamedSelection][1])
for MinNum in Faces:
	for MaxNum1 in Faces[MinNum]:
		for MidNum1 in Faces[MinNum][MaxNum1]:
			NamedSelection=BlocksList[Faces[MinNum][MaxNum1][MidNum1]]
			i=BlocksInfo[NamedSelection][3][MinNum]
			j=BlocksInfo[NamedSelection][3][MidNum1]
			k=BlocksInfo[NamedSelection][3][MaxNum1]
			BlocksInfo[NamedSelection][2].InsertNextCell(vtk.VTK_TRIANGLE,3,[i,j,k]) #triangle Other: VTK_QUAD (2D),VTK_TETRA (3D), VTK_HEXAHEDRON (3D)
#----------------------------------------------------------
#------------OUTPUT----------------------------------------
output=vtk.vtkMultiBlockDataSet()
BlocksNum=len(BlocksInfo)
output.SetNumberOfBlocks(BlocksNum)
i=0
for NamedSelection in BlocksInfo:
	if NamedSelection=='':AnnotationText=''
	elif float(ws.cell(TableRows[NamedSelection],12).value)!=-1e-30:
		Temp=float(ws.cell(TableRows[NamedSelection],12).value)
		Press=float(ws.cell(TableRows[NamedSelection],9).value)
		Array1=vtk.vtkFloatArray()
		Array1.SetName('Temp')
		Array3=vtk.vtkFloatArray()
		Array3.SetName('Press')
		for j in BlocksInfo[NamedSelection][3]:
			Array1.InsertNextValue(Temp)
			Array3.InsertNextValue(Press)
		BlocksInfo[NamedSelection][2].GetPointData().SetScalars(Array1)
		BlocksInfo[NamedSelection][2].GetPointData().AddArray(Array3)
	elif float(ws.cell(TableRows[NamedSelection],13).value)!=-1e-30:
		HTC=float(ws.cell(TableRows[NamedSelection],13).value)
		Press=float(ws.cell(TableRows[NamedSelection],9).value)
		Array2=vtk.vtkFloatArray()
		Array2.SetName('Heat Flux')
		Array3=vtk.vtkFloatArray()
		Array3.SetName('Press')
		for j in BlocksInfo[NamedSelection][3]:
			Array2.InsertNextValue(HTC)
			Array3.InsertNextValue(Press)
		BlocksInfo[NamedSelection][2].GetPointData().SetScalars(Array2)
		BlocksInfo[NamedSelection][2].GetPointData().AddArray(Array3)
	else:
		Temp=float(ws.cell(TableRows[NamedSelection],7).value)
		HTC=float(ws.cell(TableRows[NamedSelection],8).value)
		Press=float(ws.cell(TableRows[NamedSelection],9).value)
		Array1=vtk.vtkFloatArray()
		Array1.SetName('Temp sink')
		Array2=vtk.vtkFloatArray()
		Array2.SetName('HTC')
		Array3=vtk.vtkFloatArray()
		Array3.SetName('Press')
		for j in BlocksInfo[NamedSelection][3]:
			Array1.InsertNextValue(Temp)
			Array2.InsertNextValue(HTC)
			Array3.InsertNextValue(Press)
		BlocksInfo[NamedSelection][2].GetPointData().SetScalars(Array1)
		BlocksInfo[NamedSelection][2].GetPointData().AddArray(Array2)
		BlocksInfo[NamedSelection][2].GetPointData().AddArray(Array3)
	output.GetMetaData(i).Set(vtk.vtkCompositeDataSet.NAME(), NamedSelection)
	output.SetBlock(i, BlocksInfo[NamedSelection][2])
	i+=1
wb.close()
print('***'+str(BlocksNum)+' blocks output')
if BlocksNum!=0:
	MBDSwriter=vtk.vtkXMLMultiBlockDataWriter()
	MBDSwriter.SetFileName(Dir+MeshFileName[0:-3]+'vtm')
	MBDSwriter.SetDataModeToBinary()
	MBDSwriter.SetInputData(output)
	MBDSwriter.Write()
